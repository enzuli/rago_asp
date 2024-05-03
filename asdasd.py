import pandas as pd
import re, os
from docx import Document
import caratulas
from PyPDF2 import PdfReader 
import time
from openpyxl import load_workbook
import tabula
import datetime
import xlrd as xlrd


def read(path, cols, names,header, sheetname) -> pd.DataFrame:
    return pd.read_excel(path,
            usecols=cols, 
            names=names, 
            header=header,
            sheet_name=sheetname)


def buscar(path):

    equipos = read(os.path.join(path,"equipos air liq.xlsx"), "B",["Equipo"],0,"Hoja1")
    crono = read(os.path.join(path,"equipos air liq.xlsx"), "A",["Hecho"],1,"Hoja2")
    eq = [re.sub("Tanque-","",x["Equipo"]).strip() for _,x in equipos.iterrows()]
    cc = [re.sub("ALASA Nª","",x["Hecho"]).strip() for _,x in crono.iterrows()]
    faltantes_cc = ["ALASA Nª "+str(equipo) for equipo in eq if equipo not in cc]
    faltantes_solicitud = ["ALASA Nª "+str(equipo) for equipo in cc if equipo not in eq]
    if len(faltantes_cc) > len(faltantes_solicitud):
        for i in range(len(faltantes_cc) - len(faltantes_solicitud)):
            faltantes_solicitud.append("")
    elif len(faltantes_cc) < len(faltantes_solicitud):
        for i in range(len(faltantes_solicitud) - len(faltantes_cc)):
            faltantes_cc.append("")
    aux = {"Faltantes solicitud": faltantes_solicitud, "Hacer CC": faltantes_cc}
    df = pd.DataFrame.from_dict(aux)
    df.to_excel(os.path.join(path,'cc.xlsx'), index=False)
    

def completar_planilla_anexa(path,nombre_f):
    df = read(os.path.join(path, nombre_f), "C,F", ["Litros", "PT"], 0, "page 1")
    aux = {"Tarea":[],"Presion": [], "Volumen":[]}
    for _,r in df.iterrows():
        aux["Tarea"].append("Inspección Periódica")
        aux["Volumen"].append(r["Litros"]/1000)
        aux["Presion"].append(r["PT"])

    df = pd.DataFrame.from_dict(aux)
    df.to_excel(os.path.join(path,'equipos air liq final.xlsx'), index=False)


def hacer_caratulas(path):
    df = read(os.path.join(path, "fs airliq.xlsx"), "A:D", ["Denominacion","Id","Vol","PT"], 0, "page 1")
    for _,r in df.iterrows():
        eq = r["Denominacion"] if len(r["Denominacion"]) < len(r["Id"]) else r["Id"]
        vol = r["Vol"]/1000
        caratulas.crear_caratula(path, 
                                 denominacion=r["Denominacion"],
                                 volumen=vol, 
                                 cliente="Pilisar SA",
                                 presion=r["PT"],
                                 interno=r["Id"],
                                 fluido="Aire",
                                 filename=r["Id"],
                                 tarea="Control Periodico",
                                 cf=False)


def actas(path,acta,index,cantidad_equipos):
    file = PdfReader(os.path.join(path,acta))
    aux = ""
    for p in file.pages:
        aux += p.extract_text()
    n_acta = aux[re.search("VERIFICACION", aux).span()[1]:re.search("Fecha:", aux).span()[0]].strip()
    fecha = aux[re.search("Fecha:", aux).span()[1]:re.search("Habilitación", aux).span()[0]].strip()
    fecha = fecha[re.match("[0-9]{2}/[0-9]{2}/[0-9]{4}",fecha).span()[0]:re.match("[0-9]{2}/[0-9]{2}/[0-9]{4}",fecha).span()[1]]
    tarea = aux[re.search("Fecha:", aux).span()[1]:re.search("Util", aux).span()[1]].strip()
    tarea = tarea[re.search(f"{fecha}", tarea).span()[1]:re.search("Util", tarea).span()[1]]
    tarea = tarea[re.search(f"x", tarea).span()[1]:re.search("Util", tarea).span()[1]].strip().split()[0]
    if tarea == "Renovación":
        tarea = "Control Periodico"
    elif tarea == "Ext.Vida":
        tarea = "Extension de Vida Util"
    comitente = aux[re.search("Establecimiento:", aux).span()[1]:re.search("Rubro", aux).span()[0]].strip().replace("\n", " ")
    fabricante = aux[re.search("Marca:", aux).span()[1]:re.search("Modelo", aux).span()[0]].strip().replace("\n", " ")
    modelo = aux[re.search("Modelo:", aux).span()[1]:re.search("Año", aux).span()[0]].strip().replace("\n", " ")
    id = aux[re.search("Identificacion Interna:", aux).span()[1]:re.search("Registro", aux).span()[0]].strip().replace("\n", " ")
    año_de_fabricacion = aux[re.search("Año de Fabricación:", aux).span()[1]:re.search("Identificacion", aux).span()[0]].strip()
    fluido = aux[re.search("Fluido Contenido:", aux).span()[1]:re.search("Capacidad", aux).span()[0]].strip()
    volumen = aux[re.search("Capacidad:", aux).span()[1]:re.search("Fluido Refrigerante", aux).span()[0]].strip()
    norma = aux[re.search("construcción:", aux).span()[1]:re.search("Temperatura de diseño:", aux).span()[0]].strip()
    material = aux[re.search("Material:", aux).span()[1]:re.search("Presión", aux).span()[0]].strip()
    presion = aux[re.search("Presión de Trabajo", aux).span()[1]:re.search("Camisa:", aux).span()[0]].replace("Cuerpo:", "").replace("Kg/cm2","").strip()
    temperatura_trabajo = aux[re.search("Temperatura de trabajo:", aux).span()[1]:re.search("Material", aux).span()[0]].strip()
    envolvente = aux[re.search("Envolvente", aux).span()[1]:re.search("Cabezales", aux).span()[0]].strip()
    diametro_envolvente = envolvente[re.search("Diámetro:", envolvente).span()[1]:re.search("mm", envolvente).span()[0]].strip()
    longitud_envolvente = envolvente[re.search("Longitud:", envolvente).span()[1]:re.search("Superficie", envolvente).span()[0]].strip().replace("mm","").strip()
    emm_envolvente = aux[re.search("Espesor:", aux).span()[1]:re.search("Cabezales", aux).span()[0]].replace("mm.","").strip()
    cabezal= aux[re.search("Tipo de Cabezal:", aux).span()[1]:re.search("Sup. Desarrollada:", aux).span()[0]].strip()
    aislacion = aux[re.search("Aislación", aux).span()[1]:re.search("Posee", aux).span()[0]].strip()
    emm_cabezal = aux[re.search("Tipo de Cabezal:", aux).span()[1]:re.search("Sepertin", aux).span()[0]].strip()
    emm_cabezal = emm_cabezal[re.search("Derecho", emm_cabezal).span()[1]:re.search("Tubos", emm_cabezal).span()[0]].split("\n")
    emm_cabezal[0] = emm_cabezal[0][re.search("Medido:", emm_cabezal[0]).span()[1]:].replace("mm.","").strip()
    emm_cabezal[1] = emm_cabezal[1][re.search("Medido:", emm_cabezal[1]).span()[1]:].replace("mm.","").strip()
    emm_cabezal = emm_cabezal[0] if emm_cabezal[0] < emm_cabezal[1] else emm_cabezal[1]
    if "o2" in fluido.lower() or "n2" in fluido.lower() or "ar" in fluido.lower():
        dias = 365*5
    elif "co2" in fluido.lower():
        dias = 365*10
    else:
        dias = 365 
    
    date = fecha.split("/")
    date = datetime.date(int(date[2]),int(date[1]),int(date[0]))
    prox_control = (date + datetime.timedelta(days=dias)).strftime("%d/%m/%Y")
    caratulas.crear_memoria_desc(r"C:\Users\Cristian\Documents\GO",
            fecha,
            comitente,
            fabricante,
            modelo,
            id,
            año_de_fabricacion,
            fluido,
            volumen,
            norma,
            material,
            presion,
            diametro_envolvente,
            longitud_envolvente,
            emm_envolvente,
            cabezal,
            emm_cabezal,
            aislacion,
            tarea,
            temperatura_trabajo,
            n_acta,
            prox_control,
            index,
            cantidad_equipos)

    
def read_actas(path):
    pre = time.time()
    for c,f in enumerate(os.listdir(path)):
        print(f)
        actas(path,f,c,len(os.listdir(path)))
    elapsed = time.time() - pre
    print(elapsed)


def info_clientes(path):
    aux = {"Cliente":[],
           "Calle": [],
           "Nº":[],
           "Partido":[],
           "Localidad":[],
           "CP":[]}
    for f in os.listdir(path):
        wb = load_workbook(os.path.join(path,f))
        hoja = wb.active
        aux["Cliente"].append(hoja["G4"].value)
        aux["Calle"].append(hoja["G5"].value)
        aux["Localidad"].append(hoja["G6"].value)
        aux["Partido"].append(hoja["G7"].value)
        aux["Nº"].append(hoja["N5"].value)
        aux["CP"].append(hoja["N6"].value)
        
    out = pd.DataFrame.from_dict(aux)
    out.to_excel(os.path.join(path,f'Templates\Clientes Ubicacion.xlsx'), index=False)
   

def leer_memorias(path):
    out = {"Norma":[], "Material":[], "Tension":[]}
    for f in os.listdir(path):
        if f.endswith(".xlsx"):
            memoria = load_workbook(os.path.join(path,f)).active
            norma = memoria["M26"].value.lower() if type(memoria["M26"].value) == str else memoria["M26"].value
            material =memoria["M25"].value.lower() if type(memoria["M25"].value) == str else memoria["M25"].value
            out["Norma"].append(norma)
            out["Material"].append(material)
            if "asme" in str(norma):
                out["Tension"].append(memoria["AE15"].value)
                print(norma, material, memoria["AE15"].value)
            else:
                out["Tension"].append(memoria["AE14"].value)
                print(norma, material, memoria["AE14"].value)

    df = pd.DataFrame.from_dict(out)
    df.to_excel(os.path.join(path,"Tensiones.xlsx"))

read_actas(r"C:\Users\Cristian\Documents\GO\actas")

